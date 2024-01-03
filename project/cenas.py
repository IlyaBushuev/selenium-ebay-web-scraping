import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import numpy as np
from openpyxl import load_workbook 
from datetime import datetime
import time

usr_input = input("Ievadiet preces nosaukumu, lai iegūtu vidējo cenu: ")

# Set up Selenium WebDriver
service = Service()
option = webdriver.ChromeOptions()
driver = webdriver.Chrome(service=service, options=option)

# Open the required url
url = "https://www.ebay.com/sch/ebayadvsearch"
driver.get(url)

# Find and input the search box
find_input = driver.find_element(By.ID, '_nkw')
find_input.send_keys(usr_input)

# Select the item condition (new)
button_condition = driver.find_element(By.ID, "s0-1-17-6[4]-[0]-LH_ItemCondition")
button_condition.click()

# Select the buying format (buy now)
button_buying_format = driver.find_element(By.ID, "s0-1-17-6[3]-[2]-LH_BIN")
button_buying_format.click()

# Pressing ENTER and letting the page to load for 2 seconds
find_input.send_keys(Keys.ENTER)
time.sleep(2)

# Extract prices from the search results
price_elements = driver.find_elements(By.CLASS_NAME, "s-item__price")
prices = []
for element in price_elements:
    if element.text == "" or "to" in element.text:
        continue
    else:
        prices.append(float(element.text[1:].replace("," , "")))

    # Calculate the first and third quartiles (Q1 and Q3)
    q1 = np.percentile(prices, 25)
    q3 = np.percentile(prices, 75)

    # Calculate the interquartile range (IQR)
    iqr = q3 - q1

    # Define lower and upper bounds to identify outliers
    lower_bound = q1 - 1.5 * iqr
    upper_bound = q3 + 1.5 * iqr

    # Remove outliers
    filtered_data = [value for value in prices if lower_bound <= value <= upper_bound]

# Find the average price of the items in the list
average_price = np.mean(sorted(filtered_data))
print(average_price.round(2))

# Opening xlsx file and writing extracted data
wb = load_workbook("price.xlsx")
ws = wb.active

next_row = ws.max_row + 1
current_date = datetime.now().strftime("%d-%m-%Y")

ws['A' + str(next_row)].value=usr_input
ws['B' + str(next_row)].value=current_date
ws['C' + str(next_row)].value=average_price.round(2)

# Save and close the Excel file
wb.save('price.xlsx')
wb.close()